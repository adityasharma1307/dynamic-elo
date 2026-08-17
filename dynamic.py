import os
import re
import logging
import warnings
from collections import deque

import numpy as np
import pandas as pd
import xgboost as xgb
from sklearn.metrics import (
    accuracy_score, log_loss, roc_auc_score, classification_report
)
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
ATP_CLONE_DIR = os.path.join(BASE_DIR, "data", "tennis_atp")
GS_PARQUET    = os.path.join(BASE_DIR, "data", "cleaned", "atp_grand_slams_1990_2025.parquet")
CM_PARQUET    = os.path.join(BASE_DIR, "data", "cleaned", "charting_matches_clean.parquet")
CS_PARQUET    = os.path.join(BASE_DIR, "data", "cleaned", "charting_stats_clean.parquet")
OUTPUT_DIR    = os.path.join(BASE_DIR, "data", "reports")

SEED = 42

SPLITS = {
    "Split A (Test=2024)": dict(train_end=2022, tune_year=2023, test_year=2024),
    "Split B (Test=2025)": dict(train_end=2023, tune_year=2024, test_year=2025),
}

ELO_START = 1500
LEVEL_K   = {"G": 1.00, "M": 0.85, "A": 0.75, "F": 0.90, "D": 0.70, "O": 0.80}
ROUND_K   = {"F": 1.00, "SF": 0.90, "QF": 0.85, "R16": 0.80, "R32": 0.78,
             "R64": 0.75, "R128": 0.72, "RR": 0.85, "BR": 0.75}

ROLL          = 20
ROLL_BP       = 20
ROLL_FORM     = 10
RECENCY_DECAY = 0.92
FORM_DECAY    = 0.85
MIN_SURF      = 15
ATP_H         = 185

POP_BP  = 0.42
POP_SWR = 0.64
POP_RPW = 0.36

XGB_PARAMS = dict(
    n_estimators=1000,
    learning_rate=0.03,
    max_depth=4,
    subsample=0.8,
    colsample_bytree=0.8,
    min_child_weight=5,
    gamma=0.05,
    reg_lambda=1.5,
    eval_metric="logloss",
    early_stopping_rounds=40,
    random_state=SEED,
    verbosity=0,
)
SURFACE_WEIGHT = 0.75

DYNAMIC_FEATURES = [
    "elo_diff",
    "welo_diff",
    "elo_blend_diff",
    "welo_blend_diff",
    "d_swr",
    "d_sswr",
    "d_rpw",
    "d_rank",
    "F",
    "C",
    "B",
]

SLAM_SURF = {
    "Australian Open": "Hard",
    "Roland Garros":   "Clay",
    "Wimbledon":       "Grass",
    "US Open":         "Hard",
}


def parse_games(score):
    sets = str(score).strip().split()
    wg, lg = 0, 0
    for s in sets:
        s_clean = re.sub(r"\(\d+\)", "", s)
        if "-" in s_clean:
            try:
                p = s_clean.split("-")
                wg += int(p[0])
                lg += int(p[1])
            except Exception:
                pass
    return wg, lg


def load_all_atp():
    frames = []
    for yr in range(1990, 2026):
        fp = os.path.join(ATP_CLONE_DIR, f"atp_matches_{yr}.csv")
        if os.path.exists(fp):
            frames.append(pd.read_csv(fp, low_memory=False))
    if not frames:
        raise FileNotFoundError(
            f"No ATP CSVs found in {ATP_CLONE_DIR}. "
            f"Please clone https://github.com/JeffSackmann/tennis_atp.git there."
        )
    df = pd.concat(frames, ignore_index=True)

    df["tourney_date"] = pd.to_datetime(
        df["tourney_date"].astype(str), format="%Y%m%d", errors="coerce"
    )
    df["year"] = df["tourney_date"].dt.year

    for col in ["winner_name", "loser_name"]:
        df[col] = df[col].astype(str).str.lower().str.strip()

    hmap = {"R": "R", "L": "L", "U": "Unknown", "nan": "Unknown", "": "Unknown"}
    for col in ["winner_hand", "loser_hand"]:
        if col in df.columns:
            df[col] = (
                df[col].astype(str).str.strip().map(hmap).fillna("Unknown")
            )

    num = [
        "winner_rank", "loser_rank", "winner_ht", "loser_ht", "best_of",
        "minutes",
        "w_1stIn", "w_1stWon", "w_svpt", "w_2ndWon",
        "l_1stIn", "l_1stWon", "l_svpt", "l_2ndWon",
        "w_bpFaced", "w_bpSaved", "l_bpFaced", "l_bpSaved",
    ]
    for c in num:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    df = df.dropna(subset=["winner_name", "loser_name", "tourney_date"])
    df = df[~df["score"].astype(str).str.contains(r"W/O|RET|DEF", na=False)]
    df = df.sort_values(["tourney_date", "match_num"]).reset_index(drop=True)

    games = df["score"].apply(parse_games)
    df["w_games"]  = [g[0] for g in games]
    df["l_games"]  = [g[1] for g in games]
    df["game_pct"] = df["w_games"] / (df["w_games"] + df["l_games"]).clip(1)

    df["w_swr"] = (
        (df["w_1stWon"].fillna(0) + df["w_2ndWon"].fillna(0))
        / df["w_svpt"].clip(1)
    ).clip(0, 1)
    df["l_swr"] = (
        (df["l_1stWon"].fillna(0) + df["l_2ndWon"].fillna(0))
        / df["l_svpt"].clip(1)
    ).clip(0, 1)
    df["w_rpw"] = (
        (df["l_svpt"].fillna(0) - df["l_1stWon"].fillna(0) - df["l_2ndWon"].fillna(0))
        / df["l_svpt"].clip(1)
    ).clip(0, 1)
    df["l_rpw"] = (
        (df["w_svpt"].fillna(0) - df["w_1stWon"].fillna(0) - df["w_2ndWon"].fillna(0))
        / df["w_svpt"].clip(1)
    ).clip(0, 1)

    log.info(
        "All ATP loaded: %d rows | year range %d-%d",
        len(df), int(df["year"].min()), int(df["year"].max())
    )
    return df


def load_charting_bp():
    """Return per-player break-point-conversion history as a chronologically
    sorted list of (charting_date, conversion_rate) events, NOT a pre-filled
    rolling window. Pre-filling a single deque here (as before) bakes each
    player's *entire* career (including matches played years after the ATP
    match being predicted) into one static value used for every row in
    compute_features -- a data-leakage bug. compute_features() now advances
    a per-player pointer into these sorted events as it walks matches
    chronologically, so only charting matches strictly before the current
    ATP match are ever visible.
    """
    cs = pd.read_parquet(CS_PARQUET)
    cm = pd.read_parquet(CM_PARQUET)
    pairs = (
        cs.groupby("match_id")
        .agg(
            p1=("player", "first"),
            p2=("player", "last"),
            p1_bk=("bk_pts", "first"),
            p2_bk=("bk_pts", "last"),
            p1_bs=("bp_saved", "first"),
            p2_bs=("bp_saved", "last"),
        )
        .reset_index()
    )
    pairs["p1_conv"] = (
        pairs["p1_bk"] / (pairs["p1_bk"] + pairs["p2_bs"]).clip(1)
    ).clip(0, 1)
    pairs["p2_conv"] = (
        pairs["p2_bk"] / (pairs["p2_bk"] + pairs["p1_bs"]).clip(1)
    ).clip(0, 1)
    pairs = pairs.merge(
        cm[["match_id", "charting_date"]], on="match_id", how="left"
    )
    pairs = pairs.dropna(subset=["charting_date"]).sort_values("charting_date")
    pairs["charting_date"] = pd.to_datetime(pairs["charting_date"])

    events = {}
    for r in pairs.itertuples(index=False):
        for name, conv in [(r.p1, r.p1_conv), (r.p2, r.p2_conv)]:
            if pd.notna(conv):
                events.setdefault(name, []).append(
                    (r.charting_date, float(conv))
                )
    for name in events:
        events[name].sort(key=lambda x: x[0])

    log.info("Charting BP history: %d players (event-based, leak-safe)", len(events))
    return events


def recency_weighted_mean(values, decay=RECENCY_DECAY):
    if not values:
        return np.nan
    vals = list(values)
    n = len(vals)
    weights = np.array([decay ** (n - 1 - i) for i in range(n)])
    return float(np.average(vals, weights=weights))

def add_omega_variants(gs):
    """Ω was originally a plain linear sum Ω = b1*F + b2*C + b3*B, left for
    XGBoost to weight. Here we engineer explicit non-linear combinations of
    F (fatigue), C (clutch) and B (biometric edge) -- multiplicative
    interactions, signed squares, and a signed-exponential blow-up of the
    combined signal -- as additional candidate columns. They're all
    "advantage to A" signed like F/C/B themselves, so frame_as_matchup()
    flips them too. XGBoost's feature-importance ranking (see
    print_feature_importance / the Omega feature-search experiment in
    run()) tells us empirically whether any of these non-linear encodings
    give the trees an easier signal to split on than raw F, C, B alone.

    C lives on a much larger numeric scale than F and B (roughly -100..100
    vs. -1.5..1.5), so it's rescaled to c_s = C / 100 before combining --
    otherwise any product/sum involving C would be dominated by scale, not
    by the actual interaction we're trying to expose.
    """
    F = gs["F"].astype(float)
    C = gs["C"].astype(float)
    B = gs["B"].astype(float)
    c_s = C / 100.0  # rescale onto the same rough [-1, 1] order as F, B

    def signed_pow(x, p):
        return np.sign(x) * (np.abs(x) ** p)

    # Pairwise multiplicative interactions
    gs["FxC"]   = (F * c_s).astype(float)
    gs["FxB"]   = (F * B).astype(float)
    gs["CxB"]   = (c_s * B).astype(float)
    gs["FxCxB"] = (F * c_s * B).astype(float)

    # Signed squares (preserve "advantage to A" direction, punish
    # small values less / large values more than the linear term)
    gs["F_sq"] = signed_pow(F, 2).astype(float)
    gs["C_sq"] = signed_pow(c_s, 2).astype(float)
    gs["B_sq"] = signed_pow(B, 2).astype(float)

    # Signed exponential of the combined linear signal -- an "exponentiated
    # Omega": Omega_exp = sign(Omega) * (e^|Omega| - 1), clipped so it can't
    # blow up on outliers
    omega_lin = F + c_s + B
    gs["Omega_exp"] = (
        np.sign(omega_lin) * (np.exp(omega_lin.abs().clip(upper=5)) - 1)
    ).astype(float)

    # Geometric-mean-style 3-way combination (signed cube root of the
    # product) -- a smoother, scale-robust alternative to the raw product
    gs["Omega_geom"] = signed_pow(F * c_s * B, 1.0 / 3.0).astype(float)


OMEGA_CANDIDATE_COLS = [
    "FxC", "FxB", "CxB", "FxCxB",
    "F_sq", "C_sq", "B_sq",
    "Omega_exp", "Omega_geom",
]

CORE_FEATURES = [
    "elo_diff", "welo_diff", "elo_blend_diff", "welo_blend_diff",
    "d_swr", "d_sswr", "d_rpw", "d_rank",
]

OMEGA_VARIANTS = {
    "Linear (F + C + B)":        ["F", "C", "B"],
    "+ Multiplicative (FxC,FxB,CxB,FxCxB)": ["F", "C", "B", "FxC", "FxB", "CxB", "FxCxB"],
    "+ Signed-squared (F^2,C^2,B^2)":       ["F", "C", "B", "F_sq", "C_sq", "B_sq"],
    "+ Signed-exponential (Omega_exp)":     ["F", "C", "B", "Omega_exp"],
    "+ Geometric 3-way (Omega_geom)":       ["F", "C", "B", "Omega_geom"],
    "+ All combinations":        ["F", "C", "B"] + OMEGA_CANDIDATE_COLS,
}


def compute_features(df_all, bp_events):
    log.info("Single-pass feature computation")

    elo_n  = {}
    elo_s  = {}
    welo_n = {}
    welo_s = {}
    surf_n = {}

    h_swr  = {}
    h_sswr = {}
    h_rpw  = {}
    h_form = {}
    h_bp_a = {}
    win_streak = {}

    bp_hist = {}
    bp_ptr  = {}

    def advance_bp(name, cutoff_date):
        """Reveal charting break-point events for `name` strictly before
        `cutoff_date` into the rolling window. Called just before that
        player's BP feature is read for a match, so no future charting
        data can leak into a prediction for an earlier match."""
        ev = bp_events.get(name)
        if not ev:
            return
        ptr = bp_ptr.get(name, 0)
        n = len(ev)
        while ptr < n and ev[ptr][0] < cutoff_date:
            if name not in bp_hist:
                bp_hist[name] = deque(maxlen=ROLL_BP)
            bp_hist[name].append(ev[ptr][1])
            ptr += 1
        bp_ptr[name] = ptr

    def grm(hd, k, decay=RECENCY_DECAY):
        h = hd.get(k, deque())
        return recency_weighted_mean(h, decay) if h else np.nan

    def blend(n, s, ov, sv):
        w = min(surf_n.get((n, s), 0) / MIN_SURF, 1.0)
        return w * sv + (1 - w) * ov

    cols = [
        "we", "le", "wes", "les", "wb", "lb",
        "wwe", "wle", "wwes", "wles", "wwb", "wlb",
        "wswr", "lswr", "wsswr", "lsswr", "wrpw", "lrpw",
        "wbpc", "lbpc", "wbpa", "lbpa",
        "wrank", "lrank", "wht", "lht", "whand", "lhand",
        "wwst", "lwst", "wform", "lform", "wfat", "lfat",
        "_tourney_id", "_match_num",
    ]
    out = {k: [] for k in cols}
    prev_tid = None
    t_mins = {}

    for _, row in df_all.iterrows():
        wn = row["winner_name"]
        ln = row["loser_name"]
        surf = str(row.get("surface", "Hard"))
        lev  = str(row.get("tourney_level", "A"))
        rnd  = str(row.get("round", "R32"))
        bo   = int(row.get("best_of", 3) or 3)
        is_gs = (lev == "G")

        we  = elo_n.get(wn, ELO_START)
        le  = elo_n.get(ln, ELO_START)
        wes = elo_s.get((wn, surf), ELO_START)
        les = elo_s.get((ln, surf), ELO_START)
        wwe  = welo_n.get(wn, ELO_START)
        wle  = welo_n.get(ln, ELO_START)
        wwes = welo_s.get((wn, surf), ELO_START)
        wles = welo_s.get((ln, surf), ELO_START)

        if is_gs:
            tid = row.get("tourney_id")
            if tid != prev_tid:
                t_mins = {}
                prev_tid = tid

            advance_bp(wn, row["tourney_date"])
            advance_bp(ln, row["tourney_date"])

            wb  = blend(wn, surf, we, wes)
            lb  = blend(ln, surf, le, les)
            wwb = blend(wn, surf, wwe, wwes)
            wlb = blend(ln, surf, wle, wles)

            for k, v in [
                ("we", we), ("le", le), ("wes", wes), ("les", les),
                ("wb", wb), ("lb", lb),
                ("wwe", wwe), ("wle", wle), ("wwes", wwes), ("wles", wles),
                ("wwb", wwb), ("wlb", wlb),
                ("wswr", grm(h_swr, wn)), ("lswr", grm(h_swr, ln)),
                ("wsswr", grm(h_sswr, (wn, surf))),
                ("lsswr", grm(h_sswr, (ln, surf))),
                ("wrpw", grm(h_rpw, wn)), ("lrpw", grm(h_rpw, ln)),
                ("wbpc", grm(bp_hist, wn)), ("lbpc", grm(bp_hist, ln)),
                ("wbpa", grm(h_bp_a, wn)), ("lbpa", grm(h_bp_a, ln)),
                ("wrank", row.get("winner_rank", np.nan)),
                ("lrank", row.get("loser_rank", np.nan)),
                ("wht", row.get("winner_ht", np.nan)),
                ("lht", row.get("loser_ht", np.nan)),
                ("whand", str(row.get("winner_hand", "R"))),
                ("lhand", str(row.get("loser_hand", "R"))),
                ("wwst", win_streak.get(wn, 0)),
                ("lwst", win_streak.get(ln, 0)),
                ("wform", grm(h_form, wn, FORM_DECAY)),
                ("lform", grm(h_form, ln, FORM_DECAY)),
                ("wfat", t_mins.get(wn, 0.0)),
                ("lfat", t_mins.get(ln, 0.0)),
                ("_tourney_id", row.get("tourney_id")),
                ("_match_num", row.get("match_num")),
            ]:
                out[k].append(v)

            em = float(row.get("minutes") if pd.notna(row.get("minutes")) else 95)
            t_mins[wn] = t_mins.get(wn, 0.0) + em
            t_mins[ln] = t_mins.get(ln, 0.0) + em

        k = (
            32 * LEVEL_K.get(lev, 0.75) * ROUND_K.get(rnd, 0.80)
            * (1.0 if bo == 5 else 0.90)
        )
        ks = k * 0.85
        exp_w  = 1 / (1 + 10 ** ((le - we) / 400))
        delta  = k * (1 - exp_w)
        exp_ws = 1 / (1 + 10 ** ((les - wes) / 400))
        elo_n[wn] = we + delta
        elo_n[ln] = le - delta
        elo_s[(wn, surf)] = wes + ks * (1 - exp_ws)
        elo_s[(ln, surf)] = les - ks * (1 - exp_ws)

        gp = (
            row["game_pct"]
            if pd.notna(row["game_pct"]) and row["game_pct"] > 0
            else 0.6
        )
        w_factor = 2 * gp
        kw  = k * w_factor
        kws = ks * w_factor
        exp_ww  = 1 / (1 + 10 ** ((wle - wwe) / 400))
        delta_w = kw * (1 - exp_ww)
        exp_wws = 1 / (1 + 10 ** ((wles - wwes) / 400))
        welo_n[wn] = wwe + delta_w
        welo_n[ln] = wle - delta_w
        welo_s[(wn, surf)] = wwes + kws * (1 - exp_wws)
        welo_s[(ln, surf)] = wles - kws * (1 - exp_wws)

        surf_n[(wn, surf)] = surf_n.get((wn, surf), 0) + 1
        surf_n[(ln, surf)] = surf_n.get((ln, surf), 0) + 1

        win_streak[wn] = win_streak.get(wn, 0) + 1
        win_streak[ln] = 0

        for name, sv, rv, fv in [
            (wn, row.get("w_swr", np.nan), row.get("w_rpw", np.nan), 1.0),
            (ln, row.get("l_swr", np.nan), row.get("l_rpw", np.nan), 0.0),
        ]:
            if name not in h_swr:
                h_swr[name] = deque(maxlen=ROLL)
            if (name, surf) not in h_sswr:
                h_sswr[(name, surf)] = deque(maxlen=ROLL)
            if name not in h_rpw:
                h_rpw[name] = deque(maxlen=ROLL)
            if name not in h_form:
                h_form[name] = deque(maxlen=ROLL_FORM)
            if pd.notna(sv):
                h_swr[name].append(float(sv))
                h_sswr[(name, surf)].append(float(sv))
            if pd.notna(rv):
                h_rpw[name].append(float(rv))
            h_form[name].append(fv)

        if is_gs:
            w_bp = max(
                (row.get("l_bpFaced", 0) or 0) - (row.get("l_bpSaved", 0) or 0),
                0,
            ) / max(row.get("l_bpFaced", 1) or 1, 1)
            l_bp = max(
                (row.get("w_bpFaced", 0) or 0) - (row.get("w_bpSaved", 0) or 0),
                0,
            ) / max(row.get("w_bpFaced", 1) or 1, 1)
            for name, bv in [(wn, w_bp), (ln, l_bp)]:
                if name not in h_bp_a:
                    h_bp_a[name] = deque(maxlen=ROLL)
                h_bp_a[name].append(float(bv))

    gs = pd.read_parquet(GS_PARQUET)
    gs["tourney_date"] = pd.to_datetime(gs["tourney_date"])
    gs["year"] = gs["tourney_date"].dt.year
    gs = gs.sort_values(["tourney_date", "match_num"]).reset_index(drop=True)
    
    # FIX: Standardize US Open naming
    gs["tourney_name"] = gs["tourney_name"].replace("Us Open", "US Open")

    # SAFETY CHECK: `out` (built by walking df_all in chronological order)
    # and `gs` (loaded independently from GS_PARQUET) are stitched together
    # by row position. That's only safe if every Grand Slam match present
    # in one is present in the other, in the same order. Verify this on the
    # overlapping prefix using tourney_id + match_num as the join key rather
    # than trusting the row counts alone -- a silent positional mismatch
    # would attach the wrong player's feature values to a match.
    n_features = len(out["we"])
    n_check = min(len(gs), n_features)
    gs_tid = gs["tourney_id"].values[:n_check]
    gs_mn  = gs["match_num"].values[:n_check]
    out_tid = np.array(out["_tourney_id"][:n_check])
    out_mn  = np.array(out["_match_num"][:n_check])
    mismatches = np.where((gs_tid != out_tid) | (gs_mn != out_mn))[0]
    if len(mismatches):
        i = mismatches[0]
        raise ValueError(
            f"Feature/row alignment mismatch at position {i}: "
            f"gs=({gs_tid[i]}, {gs_mn[i]}) vs computed=({out_tid[i]}, {out_mn[i]}). "
            "The positional join between df_all-derived features and the GS "
            "parquet is broken -- check for match filtering differences "
            "between the two sources before trusting any downstream results."
        )

    if len(gs) > n_features:
        log.warning(
            "GS has more rows (%d) than computed features (%d) -- this is "
            "expected only for rows chronologically past the end of the "
            "loaded ATP CSV history (e.g. a future season not yet cloned "
            "locally); padding with NaN.",
            len(gs), n_features,
        )
        for k in out:
            out[k] = out[k] + [np.nan] * (len(gs) - n_features)
    elif len(gs) < n_features:
        n = len(gs)
        gs = gs.iloc[:n].reset_index(drop=True)
        for k in out:
            out[k] = out[k][:n]

    for col, vals in out.items():
        if col in ("_tourney_id", "_match_num"):
            continue
        gs[col] = vals

    gs["elo_diff"]        = (gs["we"]   - gs["le"]).astype(float)
    gs["elo_blend_diff"]  = (gs["wb"]   - gs["lb"]).astype(float)
    gs["welo_diff"]       = (gs["wwe"]  - gs["wle"]).astype(float)
    gs["welo_blend_diff"] = (gs["wwb"]  - gs["wlb"]).astype(float)
    gs["d_bp"] = (
        gs["wbpc"].fillna(gs["wbpa"]).fillna(POP_BP)
        - gs["lbpc"].fillna(gs["lbpa"]).fillna(POP_BP)
    ).astype(float)
    gs["d_swr"]  = (
        gs["wswr"].fillna(POP_SWR) - gs["lswr"].fillna(POP_SWR)
    ).astype(float)
    gs["d_sswr"] = (
        gs["wsswr"].fillna(POP_SWR) - gs["lsswr"].fillna(POP_SWR)
    ).astype(float)
    gs["d_rpw"] = (
        gs["wrpw"].fillna(POP_RPW) - gs["lrpw"].fillna(POP_RPW)
    ).astype(float)
    gs["d_rank"] = (
        gs["wrank"].fillna(200) - gs["lrank"].fillna(200)
    ).astype(float)

    total_fat = (gs["wfat"] + gs["lfat"]).clip(lower=1)
    gs["F"] = (
        (gs["lfat"] - gs["wfat"]) / total_fat
    ).fillna(0.0).astype(float)

    gs["C"] = (gs["d_bp"] * 100).astype(float)

    wh = gs["whand"].values
    lh = gs["lhand"].values
    hand_adv = np.where(
        (wh == "L") & (lh == "R"), 1,
        np.where((wh == "R") & (lh == "L"), -1, 0),
    ).astype(float)
    gs["B"] = (
        (gs["wht"].fillna(187) - gs["lht"].fillna(187)) / ATP_H
        + hand_adv
    ).astype(float)

    add_omega_variants(gs)

    log.info(
        "Features computed. Raw ELO GS accuracy: %.1f%% | Final shape %s",
        (gs["elo_diff"] > 0).mean() * 100, gs.shape
    )
    return gs

def frame_as_matchup(gs):
    np.random.seed(SEED)
    gs = gs.copy()
    gs["flip"] = np.random.randint(2, size=len(gs))
    sign = np.where(gs["flip"] == 0, 1, -1)
    # Columns that are ODD functions of the underlying A/B-relative signal
    # (i.e. they flip sign when A and B are swapped) must be re-signed here.
    # F_sq/C_sq/B_sq are signed squares (still odd: sign(x)*x^2 negates when
    # x negates), FxCxB and Omega_geom are odd (odd*odd*odd=odd), and
    # Omega_exp is odd by the same construction as the signed squares.
    # FxC/FxB/CxB are each a product of two odd terms -> EVEN, i.e.
    # invariant to the A/B swap, so they are deliberately excluded here;
    # flipping them would introduce a sign bug.
    asym_cols = [
        "elo_diff", "elo_blend_diff",
        "welo_diff", "welo_blend_diff",
        "d_swr", "d_sswr", "d_rpw", "d_rank",
        "F", "C", "B",
        "F_sq", "C_sq", "B_sq", "FxCxB", "Omega_exp", "Omega_geom",
    ]
    for col in asym_cols:
        if col in gs.columns:
            gs[col] = (gs[col] * sign).astype(float)
    gs["target"] = np.where(gs["flip"] == 0, 1, 0)
    return gs

def train_surface_ensemble(train, tune, features, label):
    log.info("Training DYNAMIC ELO [%s] surface ensemble", label)
    models = {}
    for surf in ["Hard", "Clay", "Grass"]:
        tr = train[train["surface"] == surf]
        tu = tune[tune["surface"] == surf]
        if len(tr) < 100 or len(tu) < 20:
            continue
        m = xgb.XGBClassifier(**XGB_PARAMS)
        m.fit(
            tr[features], tr["target"],
            eval_set=[(tu[features], tu["target"])],
            verbose=False,
        )
        tu_acc = accuracy_score(tu["target"], m.predict(tu[features]))
        log.info(
            "  [%s] DYNAMIC [%s]: %d train | tune=%.1f%% | best_iter=%d",
            label, surf, len(tr), tu_acc * 100, m.best_iteration,
        )
        models[surf] = m

    gm = xgb.XGBClassifier(**XGB_PARAMS)
    gm.fit(
        train[features], train["target"],
        eval_set=[(tune[features], tune["target"])],
        verbose=False,
    )
    log.info(
        "  [%s] DYNAMIC [global]: tune=%.1f%% | best_iter=%d",
        label,
        accuracy_score(tune["target"], gm.predict(tune[features])) * 100,
        gm.best_iteration,
    )
    models["global"] = gm
    return models


def predict_ensemble(models, df, features, surf_weight=SURFACE_WEIGHT):
    gm = models["global"]
    gp = gm.predict_proba(df[features])[:, 1]
    blended = np.zeros(len(df))
    for i, (_, row) in enumerate(df.iterrows()):
        surf = str(row.get("surface", "Hard"))
        sm = models.get(surf)
        if sm is not None:
            X = pd.DataFrame(
                [row[features].astype(float).values], columns=features
            )
            sp = sm.predict_proba(X)[0, 1]
            blended[i] = surf_weight * sp + (1 - surf_weight) * gp[i]
        else:
            blended[i] = gp[i]
    return (blended >= 0.5).astype(int), blended


def evaluate(preds, proba, y, label):
    acc = accuracy_score(y, preds)
    ll  = log_loss(y, proba)
    auc = roc_auc_score(y, proba)
    print(f"\n{'-'*60}")
    print(f"  {label}")
    print(f"{'-'*60}")
    print(f"  Accuracy : {acc*100:.1f}%   Log Loss: {ll:.4f}   AUC: {auc:.4f}")
    print(classification_report(
        y, preds, target_names=["B wins", "A wins"], digits=3
    ))
    return {"accuracy": acc, "log_loss": ll, "roc_auc": auc}


def evaluate_per_slam(models, test, features, label):
    print(f"\n{'='*64}")
    print(f"  DYNAMIC ELO [{label}] - GRAND SLAMS BREAKDOWN")
    print(f"{'='*64}")
    summary = []
    for slam in sorted(test["tourney_name"].unique()):
        sub = test[test["tourney_name"] == slam]
        if sub.empty:
            continue
        p, pr = predict_ensemble(models, sub, features)
        acc = accuracy_score(sub["target"], p)
        auc = (
            roc_auc_score(sub["target"], pr)
            if len(sub["target"].unique()) > 1 else float("nan")
        )
        nc = int((p == sub["target"].values).sum())
        nm = len(sub)
        surf = SLAM_SURF.get(slam, "?")
        print(f"\n  {slam} [{surf}] ({nm} matches)")
        print(f"  Correct: {nc}/{nm} ({acc*100:.1f}%)   AUC: {auc:.4f}")
        summary.append({
            "Slam": slam, "Surface": surf, "Matches": nm,
            "Correct": nc, "Accuracy (%)": round(acc * 100, 1),
            "AUC": round(auc, 4),
        })
    all_p, all_pr = predict_ensemble(models, test, features)
    oa = accuracy_score(test["target"], all_p)
    auc = roc_auc_score(test["target"], all_pr)
    print(f"\n{'='*64}")
    print(f"  OVERALL : {oa*100:.1f}%   AUC: {auc:.4f}   ({len(test)} matches)")
    print(f"{'='*64}")
    return pd.DataFrame(summary), oa, auc


def print_feature_importance(models, features, label):
    fi_label = {
        "elo_diff":        "R_Diff (FiveThirtyEight ELO)",
        "welo_diff":       "WElo (Angelini 2022)",
        "elo_blend_diff":  "Blended surface ELO",
        "welo_blend_diff": "Blended WElo",
        "d_swr":           "Serve Win Ratio (delta)",
        "d_sswr":          "Surface SWR (delta)",
        "d_rpw":           "Return Pts Won (delta)",
        "d_rank":          "ATP Rank (delta)",
        "F":               "beta1*F - Fatigue Index",
        "C":               "beta2*C - Clutch Factor",
        "B":               "beta3*B - Biometric Edge",
        "FxC":             "F x C (fatigue-clutch interaction)",
        "FxB":             "F x B (fatigue-biometric interaction)",
        "CxB":             "C x B (clutch-biometric interaction)",
        "FxCxB":           "F x C x B (3-way interaction)",
        "F_sq":            "signed F^2",
        "C_sq":            "signed C^2",
        "B_sq":            "signed B^2",
        "Omega_exp":       "signed exp(|F+C+B|) - 1",
        "Omega_geom":      "signed cube-root(F*C*B)",
    }
    print(f"\n  DYNAMIC ELO [{label}] Feature Importances:")
    print(f"  {'Feature':<32} {'Gain':>8}  Bar")
    print("  " + "-" * 72)
    fi = (
        pd.Series(
            models["global"].feature_importances_, index=features
        )
        .sort_values(ascending=False)
    )
    for f, v in fi.items():
        print(
            f"  {fi_label.get(f, f):<32} {v:>8.4f}  {'|' * int(v * 70)}"
        )


def export_excel(model, test, path, features, label):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    COLORS = {
        "Australian Open": "1565C0",
        "Roland Garros":   "B71C1C",
        "Wimbledon":       "1B5E20",
        "Us Open":         "E65100",
        "US Open":         "E65100",
    }
    THIN = Border(*[Side(style="thin")] * 4)
    CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    bp, pb = predict_ensemble(model, test, features)
    test = test.copy()
    test["pred"] = bp
    test["prob"] = pb
    test["correct"] = (bp == test["target"].values)

    # NOTE: `winner_name`/`loser_name` on `test` (passed in as `test2` from
    # run_one_split) always hold the TRUE match winner/loser regardless of
    # the random A/B flip applied in frame_as_matchup. `target` tells us
    # which slot (A=1 or B=0) the true winner occupies for this row, so the
    # model's predicted winner is `winner_name` when pred==target (i.e. the
    # model picked whichever slot actually won) and `loser_name` otherwise.
    # Filtering to target==1 here would silently drop ~50% of test matches
    # (the ones where the coin flip put the true winner in slot B).
    log_rows = []
    for (_, r), pred, prob, correct in zip(
        test.iterrows(),
        test["pred"].values,
        test["prob"].values,
        test["correct"].values,
    ):
        elo_diff_true = float(r["elo_diff"]) if r["target"] == 1 else -float(r["elo_diff"])
        log_rows.append({
            "Tournament": r["tourney_name"],
            "Date": (
                r["tourney_date"].strftime("%Y-%m-%d")
                if pd.notna(r.get("tourney_date")) else ""
            ),
            "Round":   r.get("round", "?"),
            "Surface": r.get("surface", "?"),
            "Winner":  str(r.get("winner_name", "")).title(),
            "Loser":   str(r.get("loser_name", "")).title(),
            "Delta ELO":        round(float(r["elo_diff"]), 1),
            "Delta WElo":       round(float(r["welo_diff"]), 1),
            "Delta Blend ELO":  round(float(r["elo_blend_diff"]), 1),
            "Delta Blend WElo": round(float(r["welo_blend_diff"]), 1),
            "Delta Serve Win":  round(float(r["d_swr"]), 4),
            "Delta Surf SWR":   round(float(r["d_sswr"]), 4),
            "Delta Ret Pts":    round(float(r["d_rpw"]), 4),
            "Delta Rank":       round(float(r["d_rank"]), 0),
            "F (Fatigue)":      round(float(r["F"]), 4),
            "C (Clutch)":       round(float(r["C"]), 2),
            "B (Biometric)":    round(float(r["B"]), 4),
            "Win Probability":  round(prob, 4),
            "Confidence":       round(abs(prob - 0.5) * 2, 4),
            "Prediction": (
                f"{r['winner_name'].title()} wins" if pred == r["target"]
                else f"{r['loser_name'].title()} wins"
            ),
            "Correct?": "YES" if correct else "NO",
            "Upset?":   "YES" if elo_diff_true < -50 else "-",
        })
    log_df = pd.DataFrame(log_rows)

    slam_order = ["Australian Open", "Roland Garros", "Wimbledon", "US Open"]
    s_rows = []
    for slam in slam_order:
        sub = log_df[log_df["Tournament"] == slam]
        if sub.empty:
            continue
        nm = len(sub)
        nc = sub["Correct?"].str.contains("YES").sum()
        s_rows.append({
            "Tournament":   slam,
            "Surface":      SLAM_SURF.get(slam, "?"),
            "Matches":      nm,
            "Correct":      int(nc),
            "Wrong":        int(nm - nc),
            "Accuracy (%)": round(nc / nm * 100, 1),
            "Upsets":       int(sub["Upset?"].str.contains("YES").sum()),
            "Avg Confidence": round(sub["Confidence"].mean(), 4),
        })
    if s_rows:
        tn = sum(r["Matches"] for r in s_rows)
        tc = sum(r["Correct"] for r in s_rows)
        s_rows.append({
            "Tournament":   "OVERALL",
            "Surface":      "-",
            "Matches":      tn,
            "Correct":      tc,
            "Wrong":        tn - tc,
            "Accuracy (%)": round(tc / tn * 100, 1),
            "Upsets":       sum(r["Upsets"] for r in s_rows),
            "Avg Confidence": round(log_df["Confidence"].mean(), 4),
        })
    sum_df = pd.DataFrame(s_rows)

    with pd.ExcelWriter(path, engine="openpyxl") as w:
        sum_df.to_excel(w, sheet_name="Summary", index=False)
        log_df.to_excel(w, sheet_name="All Matches", index=False)
        for slam in slam_order:
            sub = log_df[log_df["Tournament"] == slam]
            if not sub.empty:
                sub.to_excel(w, sheet_name=slam[:31], index=False)

    wb = load_workbook(path)
    for sn in wb.sheetnames:
        ws = wb[sn]
        col = COLORS.get(sn, "263238")
        hf = PatternFill("solid", fgColor=col)
        for cell in ws[1]:
            cell.fill = hf
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = CTR
            cell.border = THIN
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        cc = uc = None
        for i, c in enumerate(ws[1], 1):
            if c.value == "Correct?":
                cc = i
            if c.value == "Upset?":
                uc = i
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = CTR
                cell.border = THIN
            if cc:
                f = (
                    PatternFill("solid", fgColor="C8E6C9")
                    if "YES" in str(row[cc - 1].value)
                    else PatternFill("solid", fgColor="FFCDD2")
                )
                for cell in row:
                    cell.fill = f
            if uc and "YES" in str(row[uc - 1].value):
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="FFF9C4")

        if sn == "Summary":
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                sv = str(row[0].value)
                is_ov = sv == "OVERALL"
                fc = "FFD600" if is_ov else COLORS.get(sv, "263238")
                txt = "000000" if is_ov else "FFFFFF"
                fill = PatternFill("solid", fgColor=fc)
                for cell in row:
                    cell.fill = fill
                    cell.font = Font(bold=True, size=11, color=txt)
                    cell.alignment = CTR
                    cell.border = THIN
        for col_c in ws.columns:
            w_ = min(
                max(
                    max(
                        len(str(c.value)) if c.value else 0
                        for c in col_c
                    ) + 2,
                    10,
                ),
                36,
            )
            ws.column_dimensions[
                get_column_letter(col_c[0].column)
            ].width = w_
    wb.save(path)
    log.info("[%s] Excel saved: %s", label, path)


def run_one_split(matchup, gs, label, train_end, tune_year, test_year, features=None):
    features = features or DYNAMIC_FEATURES
    print("\n" + "#" * 70)
    print(f"#  RUNNING: {label}")
    print(f"#  Train <= {train_end}  |  Tune = {tune_year}  |  Test = {test_year}")
    print(f"#  Features: {features}")
    print("#" * 70)

    train = matchup[matchup["year"] <= train_end]
    tune  = matchup[matchup["year"] == tune_year]
    test  = matchup[matchup["year"] == test_year]
    log.info(
        "[%s] Splits | Train=%d | Tune=%d | Test=%d",
        label, len(train), len(tune), len(test),
    )

    if len(test) == 0:
        log.error(
            "[%s] No test data for year %d. Aborting this split.",
            label, test_year,
        )
        return None

    models = train_surface_ensemble(train, tune, features, label)

    tune_p, tune_pr = predict_ensemble(models, tune, features)
    tune_metrics = evaluate(
        tune_p, tune_pr, tune["target"],
        f"DYNAMIC ELO [{label}] - Tune ({tune_year})",
    )

    test_p, test_pr = predict_ensemble(models, test, features)
    test_metrics = evaluate(
        test_p, test_pr, test["target"],
        f"DYNAMIC ELO [{label}] - Test ({test_year} FULL)",
    )

    slam_summary, oa, oauc = evaluate_per_slam(
        models, test, features, label,
    )
    print_feature_importance(models, features, label)

    test2 = test.copy()
    orig = gs.sort_values(
        ["tourney_date", "match_num"]
    ).reset_index(drop=True)
    test2["winner_name"] = np.where(
        test2["flip"] == 0,
        orig.loc[test2.index, "winner_name"].values,
        orig.loc[test2.index, "loser_name"].values,
    )
    test2["loser_name"] = np.where(
        test2["flip"] == 0,
        orig.loc[test2.index, "loser_name"].values,
        orig.loc[test2.index, "winner_name"].values,
    )
    test2["tourney_date"] = orig.loc[test2.index, "tourney_date"].values

    out = os.path.join(
        OUTPUT_DIR,
        f"{test_year}_GS_Dynamic_ELO_Match_Log.xlsx",
    )
    export_excel(models, test2, out, features, label)

    fi = pd.Series(
        models["global"].feature_importances_, index=features
    ).sort_values(ascending=False)

    return {
        "label": label,
        "train_end": train_end,
        "tune_year": tune_year,
        "test_year": test_year,
        "tune_metrics": tune_metrics,
        "test_metrics": test_metrics,
        "slam_summary": slam_summary,
        "overall_acc": oa,
        "overall_auc": oauc,
        "n_train": len(train),
        "n_tune": len(tune),
        "n_test": len(test),
        "feature_importance": fi,
        "features": features,
    }


def print_side_by_side(results_a, results_b):
    if results_a is None or results_b is None:
        print("\n[!] Cannot compare - one of the splits had no test data.")
        return

    print("\n" + "=" * 80)
    print("  SIDE-BY-SIDE COMPARISON - DYNAMIC ELO MODEL")
    print("=" * 80)

    la, lb = results_a["label"], results_b["label"]

    print(f"\n  {'Metric':<28} | {la:>22} | {lb:>22}")
    print("  " + "-" * 78)
    print(f"  {'Train <=':<28} | {results_a['train_end']:>22} | {results_b['train_end']:>22}")
    print(f"  {'Tune year':<28} | {results_a['tune_year']:>22} | {results_b['tune_year']:>22}")
    print(f"  {'Test year':<28} | {results_a['test_year']:>22} | {results_b['test_year']:>22}")
    print(f"  {'Train rows':<28} | {results_a['n_train']:>22} | {results_b['n_train']:>22}")
    print(f"  {'Tune rows':<28} | {results_a['n_tune']:>22} | {results_b['n_tune']:>22}")
    print(f"  {'Test rows':<28} | {results_a['n_test']:>22} | {results_b['n_test']:>22}")
    print("  " + "-" * 78)
    print(
        f"  {'Tune accuracy':<28} | "
        f"{results_a['tune_metrics']['accuracy']*100:>21.1f}% | "
        f"{results_b['tune_metrics']['accuracy']*100:>21.1f}%"
    )
    print(
        f"  {'Test accuracy':<28} | "
        f"{results_a['test_metrics']['accuracy']*100:>21.1f}% | "
        f"{results_b['test_metrics']['accuracy']*100:>21.1f}%"
    )
    print(
        f"  {'Test AUC':<28} | "
        f"{results_a['test_metrics']['roc_auc']:>22.4f} | "
        f"{results_b['test_metrics']['roc_auc']:>22.4f}"
    )
    print(
        f"  {'Test log loss':<28} | "
        f"{results_a['test_metrics']['log_loss']:>22.4f} | "
        f"{results_b['test_metrics']['log_loss']:>22.4f}"
    )

    print("\n  Per-slam test accuracy:")
    print("  " + "-" * 78)
    sa = results_a["slam_summary"].set_index("Slam")
    sb = results_b["slam_summary"].set_index("Slam")
    all_slams = sorted(set(sa.index) | set(sb.index))
    for slam in all_slams:
        a_acc = (
            f"{sa.loc[slam, 'Accuracy (%)']:>5.1f}%  ({sa.loc[slam, 'Correct']}/{sa.loc[slam, 'Matches']})"
            if slam in sa.index else "-"
        )
        b_acc = (
            f"{sb.loc[slam, 'Accuracy (%)']:>5.1f}%  ({sb.loc[slam, 'Correct']}/{sb.loc[slam, 'Matches']})"
            if slam in sb.index else "-"
        )
        print(f"  {slam:<28} | {a_acc:>22} | {b_acc:>22}")

    print("=" * 80 + "\n")


def export_comparison_excel(results_a, results_b, path, omega_results=None, best_variant=None):
    if results_a is None or results_b is None:
        return
    os.makedirs(os.path.dirname(path), exist_ok=True)

    rows = [
        ["Train <=", results_a["train_end"], results_b["train_end"]],
        ["Tune year", results_a["tune_year"], results_b["tune_year"]],
        ["Test year", results_a["test_year"], results_b["test_year"]],
        ["Train rows", results_a["n_train"], results_b["n_train"]],
        ["Tune rows", results_a["n_tune"], results_b["n_tune"]],
        ["Test rows", results_a["n_test"], results_b["n_test"]],
        [
            "Tune accuracy (%)",
            round(results_a["tune_metrics"]["accuracy"] * 100, 1),
            round(results_b["tune_metrics"]["accuracy"] * 100, 1),
        ],
        [
            "Test accuracy (%)",
            round(results_a["test_metrics"]["accuracy"] * 100, 1),
            round(results_b["test_metrics"]["accuracy"] * 100, 1),
        ],
        [
            "Test AUC",
            round(results_a["test_metrics"]["roc_auc"], 4),
            round(results_b["test_metrics"]["roc_auc"], 4),
        ],
        [
            "Test log loss",
            round(results_a["test_metrics"]["log_loss"], 4),
            round(results_b["test_metrics"]["log_loss"], 4),
        ],
    ]
    head_df = pd.DataFrame(
        rows, columns=["Metric", results_a["label"], results_b["label"]]
    )

    sa = results_a["slam_summary"].set_index("Slam")
    sb = results_b["slam_summary"].set_index("Slam")
    all_slams = sorted(set(sa.index) | set(sb.index))
    slam_rows = []
    for slam in all_slams:
        slam_rows.append({
            "Slam": slam,
            f"{results_a['label']} Accuracy (%)": (
                sa.loc[slam, "Accuracy (%)"] if slam in sa.index else None
            ),
            f"{results_a['label']} Correct/Total": (
                f"{sa.loc[slam, 'Correct']}/{sa.loc[slam, 'Matches']}"
                if slam in sa.index else "-"
            ),
            f"{results_b['label']} Accuracy (%)": (
                sb.loc[slam, "Accuracy (%)"] if slam in sb.index else None
            ),
            f"{results_b['label']} Correct/Total": (
                f"{sb.loc[slam, 'Correct']}/{sb.loc[slam, 'Matches']}"
                if slam in sb.index else "-"
            ),
        })
    slam_df = pd.DataFrame(slam_rows)

    fi_a = results_a.get("feature_importance")
    fi_df = None
    if fi_a is not None:
        fi_df = fi_a.rename("Gain").reset_index().rename(columns={"index": "Feature"})
        fi_df["Gain (%)"] = (fi_df["Gain"] * 100).round(2)

    with pd.ExcelWriter(path, engine="openpyxl") as w:
        head_df.to_excel(w, sheet_name="Overall Comparison", index=False)
        slam_df.to_excel(w, sheet_name="Per-Slam Comparison", index=False)
        if omega_results is not None:
            omega_results.to_excel(w, sheet_name="Omega Feature Search", index=False)
        if fi_df is not None:
            fi_df.to_excel(w, sheet_name="Feature Importance", index=False)

    if omega_results is not None and best_variant is not None:
        wb = load_workbook(path)
        ws = wb["Omega Feature Search"]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if str(row[0].value) == best_variant:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="C8E6C9")
                    cell.font = Font(bold=True)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="263238")
        wb.save(path)

    log.info("Comparison Excel saved: %s", path)


def run_omega_experiment(matchup, split_label="Split A (Test=2024)"):
    """Screen the Omega variants defined in OMEGA_VARIANTS (linear sum vs.
    multiplicative interactions, signed squares, signed exponential, and a
    geometric 3-way combination) on a single chronological split. Uses one
    global XGBoost model per variant (no per-surface ensemble) since this
    is a feature-selection pass, not the final production model. Returns a
    ranked results DataFrame and the name of the winning variant.
    """
    cfg = SPLITS[split_label]
    train = matchup[matchup["year"] <= cfg["train_end"]]
    tune  = matchup[matchup["year"] == cfg["tune_year"]]
    test  = matchup[matchup["year"] == cfg["test_year"]]

    print("\n" + "=" * 70)
    print("  OMEGA FEATURE-COMBINATION SEARCH")
    print("  Omega = b1*F + b2*C + b3*B (linear) vs. multiplicative /")
    print("  signed-squared / signed-exponential / geometric variants")
    print("=" * 70)

    rows = []
    for name, extra in OMEGA_VARIANTS.items():
        feats = CORE_FEATURES + extra
        m = xgb.XGBClassifier(**XGB_PARAMS)
        m.fit(
            train[feats], train["target"],
            eval_set=[(tune[feats], tune["target"])],
            verbose=False,
        )
        pred  = m.predict(test[feats])
        proba = m.predict_proba(test[feats])[:, 1]
        acc = accuracy_score(test["target"], pred)
        auc = roc_auc_score(test["target"], proba)
        ll  = log_loss(test["target"], proba)
        rows.append({
            "Omega Variant":     name,
            "Extra Features":    ", ".join(extra),
            "Test Accuracy (%)": round(acc * 100, 2),
            "Test AUC":          round(auc, 4),
            "Test Log Loss":     round(ll, 4),
        })
        print(f"  {name:<42} acc={acc*100:5.2f}%  auc={auc:.4f}  logloss={ll:.4f}")

    df = (
        pd.DataFrame(rows)
        .sort_values("Test Accuracy (%)", ascending=False)
        .reset_index(drop=True)
    )
    best = df.iloc[0]
    print(
        f"\n  BEST OMEGA VARIANT: {best['Omega Variant']} "
        f"({best['Test Accuracy (%)']:.2f}% test accuracy, "
        f"AUC {best['Test AUC']:.4f})"
    )
    print("=" * 70)
    return df, best["Omega Variant"]


def run():
    print("\n" + "=" * 70)
    print("  DYNAMIC ELO MODEL - TWO-SPLIT COMPARISON")
    print("  P_A = 1 / (1 + 10^(-(R_Diff + Omega)/400))")
    print("  Omega variants searched: linear / multiplicative / signed-")
    print("  squared / signed-exponential / geometric (see OMEGA_VARIANTS)")
    print("=" * 70)

    df_all  = load_all_atp()
    bp_hist = load_charting_bp()
    gs      = compute_features(df_all, bp_hist)
    matchup = frame_as_matchup(gs)

    omega_results, best_variant = run_omega_experiment(matchup)
    final_features = CORE_FEATURES + OMEGA_VARIANTS[best_variant]
    log.info(
        "Selected Omega variant '%s' -> final feature set: %s",
        best_variant, final_features,
    )

    results = {}
    for label, cfg in SPLITS.items():
        results[label] = run_one_split(
            matchup, gs, label,
            cfg["train_end"], cfg["tune_year"], cfg["test_year"],
            features=final_features,
        )

    keys = list(results.keys())
    print_side_by_side(results[keys[0]], results[keys[1]])

    export_comparison_excel(
        results[keys[0]],
        results[keys[1]],
        os.path.join(OUTPUT_DIR, "Dynamic_ELO_Split_Comparison.xlsx"),
        omega_results=omega_results,
        best_variant=best_variant,
    )

    return results


if __name__ == "__main__":
    results = run()