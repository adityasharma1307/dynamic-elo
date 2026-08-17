"""
Build a single self-contained Excel dashboard (data/reports/Dynamic_ELO_Dashboard.xlsx)
summarizing the Baseline vs. Dynamic ELO results for a GitHub-release showcase.

Reads from the reports already produced by baseline.py and dynamic.py:
    data/reports/2024_GS_Baseline_Match_Log.xlsx
    data/reports/2024_GS_Dynamic_ELO_Match_Log.xlsx
    data/reports/Dynamic_ELO_Split_Comparison.xlsx   (Omega Feature Search + Feature Importance)

Run baseline.py and dynamic.py first (in that order) to refresh those source
files, then run this script to refresh the dashboard.
"""
import os

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REPORTS  = os.path.join(BASE_DIR, "data", "reports")

BASELINE_LOG    = os.path.join(REPORTS, "2024_GS_Baseline_Match_Log.xlsx")
DYNAMIC_LOG     = os.path.join(REPORTS, "2024_GS_Dynamic_ELO_Match_Log.xlsx")
DYNAMIC_COMPARE = os.path.join(REPORTS, "Dynamic_ELO_Split_Comparison.xlsx")
OUT_PATH        = os.path.join(REPORTS, "Dynamic_ELO_Dashboard.xlsx")

FONT       = "Calibri"
NAVY       = "1F2937"
BLUE       = "1565C0"
GREEN      = "1B5E20"
LIGHT_GREY = "F3F4F6"
WHITE      = "FFFFFF"

THIN = Border(*[Side(style="thin", color="D1D5DB")] * 4)
CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT  = Alignment(horizontal="left", vertical="center")


def header_row(ws, row, ncols, fill=NAVY, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = CTR
        cell.border = THIN


def autosize(ws, max_col, max_row, min_w=10, max_w=42):
    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        best = min_w
        for r in range(1, max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                best = max(best, len(str(v)) + 2)
        ws.column_dimensions[col_letter].width = min(best, max_w)


def write_table(ws, df, start_row=1, start_col=1, header_fill=NAVY, number_cols=None):
    number_cols = number_cols or {}
    ncols = len(df.columns)
    for j, col in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col + j, value=col)
    header_row(ws, start_row, ncols, fill=header_fill, start_col=start_col)
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        for j, col in enumerate(df.columns):
            cell = ws.cell(row=start_row + i, column=start_col + j, value=row[col])
            cell.font = Font(name=FONT, size=10)
            cell.alignment = LFT if j == 0 else CTR
            cell.border = THIN
            if col in number_cols:
                cell.number_format = number_cols[col]
    return start_row + len(df)  # last data row


def main():
    if not (os.path.exists(BASELINE_LOG) and os.path.exists(DYNAMIC_LOG) and os.path.exists(DYNAMIC_COMPARE)):
        raise FileNotFoundError(
            "Missing source reports. Run `python baseline.py` then `python dynamic.py` "
            "first to generate the files this dashboard is built from."
        )

    base_summary = pd.read_excel(BASELINE_LOG, sheet_name="Summary")
    base_matches = pd.read_excel(BASELINE_LOG, sheet_name="All Matches")
    dyn_summary  = pd.read_excel(DYNAMIC_LOG, sheet_name="Summary")
    dyn_matches  = pd.read_excel(DYNAMIC_LOG, sheet_name="All Matches")
    omega        = pd.read_excel(DYNAMIC_COMPARE, sheet_name="Omega Feature Search")
    fi           = pd.read_excel(DYNAMIC_COMPARE, sheet_name="Feature Importance")

    wb = Workbook()
    wb.remove(wb.active)

    # ------------------------------------------------------------------ #
    # Dashboard sheet
    # ------------------------------------------------------------------ #
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    ws["B2"] = "Dynamic ELO — Tennis Grand Slam Prediction"
    ws["B2"].font = Font(name=FONT, bold=True, size=18, color=NAVY)
    ws["B3"] = "Baseline vs. Dynamic ELO | 2024 Grand Slam season | 486 matches"
    ws["B3"].font = Font(name=FONT, size=11, italic=True, color="6B7280")

    # KPI cards
    kpis = [
        ("Baseline Accuracy", f"{base_summary.loc[base_summary['Tournament']=='OVERALL','Accuracy (%)'].iloc[0]:.1f}%", BLUE),
        ("Dynamic ELO Accuracy", f"{dyn_summary.loc[dyn_summary['Tournament']=='OVERALL','Accuracy (%)'].iloc[0]:.1f}%", GREEN),
        ("Improvement", None, "B45309"),  # filled via formula below
        ("Matches Evaluated", f"{int(dyn_summary.loc[dyn_summary['Tournament']=='OVERALL','Matches'].iloc[0])}", NAVY),
    ]
    kpi_row = 5
    for i, (label, value, color) in enumerate(kpis):
        col = 2 + i * 3
        ws.merge_cells(start_row=kpi_row, start_column=col, end_row=kpi_row, end_column=col + 1)
        ws.merge_cells(start_row=kpi_row + 1, start_column=col, end_row=kpi_row + 2, end_column=col + 1)
        lbl = ws.cell(row=kpi_row, column=col, value=label)
        lbl.font = Font(name=FONT, size=10, bold=True, color=WHITE)
        lbl.fill = PatternFill("solid", fgColor=color)
        lbl.alignment = CTR
        val_cell = ws.cell(row=kpi_row + 1, column=col)
        if label == "Improvement":
            val_cell.value = "=E6-B6"  # placeholder overwritten below once numeric cells exist
        else:
            val_cell.value = value
        val_cell.font = Font(name=FONT, size=20, bold=True, color=color)
        val_cell.alignment = CTR
        for rr in (kpi_row, kpi_row + 1, kpi_row + 2):
            for cc in (col, col + 1):
                ws.cell(row=rr, column=cc).border = THIN

    # Hidden helper cells for the Improvement KPI formula (so it's a real
    # formula, not a hardcoded number)
    ws["B20"] = base_summary.loc[base_summary["Tournament"] == "OVERALL", "Accuracy (%)"].iloc[0]
    ws["B21"] = dyn_summary.loc[dyn_summary["Tournament"] == "OVERALL", "Accuracy (%)"].iloc[0]
    ws["B22"] = "=B21-B20"
    ws["A19"] = "Helper cells (used by Improvement KPI formula):"
    ws["A19"].font = Font(name=FONT, size=8, italic=True, color="9CA3AF")
    ws["A20"] = "Baseline acc (%)"
    ws["A21"] = "Dynamic acc (%)"
    ws["A22"] = "Delta (pp)"
    for r in (20, 21, 22):
        ws.cell(row=r, column=1).font = Font(name=FONT, size=8, color="9CA3AF")
        ws.cell(row=r, column=2).font = Font(name=FONT, size=8, color="9CA3AF")
        ws.cell(row=r, column=2).number_format = "0.00"
    improvement_cell = ws.cell(row=kpi_row + 1, column=2 + 2 * 3)
    improvement_cell.value = "=B22"
    improvement_cell.number_format = '+0.0"pp";-0.0"pp"'

    # --- Per-slam comparison table (source for chart 1) ---
    slam_tbl_row = 24
    ws.cell(row=slam_tbl_row, column=2, value="Per-Slam Accuracy — Baseline vs. Dynamic ELO").font = Font(
        name=FONT, bold=True, size=12, color=NAVY
    )
    slam_df = base_summary[base_summary["Tournament"] != "OVERALL"][["Tournament", "Accuracy (%)"]].rename(
        columns={"Accuracy (%)": "Baseline Accuracy (%)"}
    ).merge(
        dyn_summary[dyn_summary["Tournament"] != "OVERALL"][["Tournament", "Accuracy (%)"]].rename(
            columns={"Accuracy (%)": "Dynamic ELO Accuracy (%)"}
        ),
        on="Tournament",
    )
    last_row = write_table(
        ws, slam_df, start_row=slam_tbl_row + 1, start_col=2,
        header_fill=NAVY, number_cols={"Baseline Accuracy (%)": "0.0", "Dynamic ELO Accuracy (%)": "0.0"},
    )
    autosize(ws, 4, last_row)

    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "Per-Slam Test Accuracy (2024)"
    chart1.y_axis.title = "Accuracy (%)"
    chart1.style = 10
    data = Reference(ws, min_col=3, max_col=4, min_row=slam_tbl_row + 1, max_row=last_row)
    cats = Reference(ws, min_col=2, min_row=slam_tbl_row + 2, max_row=last_row)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.width, chart1.height = 15, 9
    ws.add_chart(chart1, "B31")

    # --- Omega feature-combination search (source for chart 2) ---
    omega_row = 24
    omega_col = 8
    ws.cell(row=omega_row, column=omega_col, value="Omega (Ω) Combination Search").font = Font(
        name=FONT, bold=True, size=12, color=NAVY
    )
    omega_small = omega[["Omega Variant", "Test Accuracy (%)"]].copy()
    last_row_o = write_table(
        ws, omega_small, start_row=omega_row + 1, start_col=omega_col,
        header_fill=BLUE, number_cols={"Test Accuracy (%)": "0.00"},
    )
    autosize(ws, omega_col + 1, last_row_o)

    chart2 = BarChart()
    chart2.type = "bar"
    chart2.title = "Omega Variant Test Accuracy"
    chart2.y_axis.title = "Accuracy (%)"
    chart2.style = 12
    data2 = Reference(ws, min_col=omega_col + 1, max_col=omega_col + 1, min_row=omega_row + 1, max_row=last_row_o)
    cats2 = Reference(ws, min_col=omega_col, min_row=omega_row + 2, max_row=last_row_o)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.width, chart2.height = 15, 9
    ws.add_chart(chart2, "H31")

    # --- Feature importance (source for chart 3) ---
    fi_row = 50
    ws.cell(row=fi_row, column=2, value="Dynamic ELO — Feature Importance (Global Model)").font = Font(
        name=FONT, bold=True, size=12, color=NAVY
    )
    fi_small = fi[["Feature", "Gain (%)"]].sort_values("Gain (%)", ascending=False).reset_index(drop=True)
    last_row_fi = write_table(
        ws, fi_small, start_row=fi_row + 1, start_col=2,
        header_fill=GREEN, number_cols={"Gain (%)": "0.00"},
    )
    autosize(ws, 3, last_row_fi)

    chart3 = BarChart()
    chart3.type = "bar"
    chart3.title = "XGBoost Feature Importance (Gain %)"
    chart3.y_axis.title = "Gain (%)"
    chart3.style = 11
    data3 = Reference(ws, min_col=3, max_col=3, min_row=fi_row + 1, max_row=last_row_fi)
    cats3 = Reference(ws, min_col=2, min_row=fi_row + 2, max_row=last_row_fi)
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)
    chart3.width, chart3.height = 15, 12
    ws.add_chart(chart3, "B67")

    ws.column_dimensions["A"].width = 2

    # ------------------------------------------------------------------ #
    # Full match-log sheets
    # ------------------------------------------------------------------ #
    for sheet_name, df in [("Baseline Matches", base_matches), ("Dynamic Matches", dyn_matches)]:
        s = wb.create_sheet(sheet_name)
        for j, col in enumerate(df.columns, start=1):
            s.cell(row=1, column=j, value=col)
        header_row(s, 1, len(df.columns))
        for i, (_, row) in enumerate(df.iterrows(), start=2):
            for j, col in enumerate(df.columns, start=1):
                cell = s.cell(row=i, column=j, value=row[col])
                cell.font = Font(name=FONT, size=9)
                cell.alignment = CTR
                if col == "Correct?":
                    cell.fill = PatternFill(
                        "solid", fgColor="C8E6C9" if row[col] == "YES" else "FFCDD2"
                    )
        s.freeze_panes = "A2"
        s.auto_filter.ref = s.dimensions
        autosize(s, len(df.columns), len(df) + 1, max_w=28)

    # ------------------------------------------------------------------ #
    # Omega Search + Feature Importance raw-data sheets
    # ------------------------------------------------------------------ #
    for sheet_name, df in [("Omega Search", omega), ("Feature Importance", fi)]:
        s = wb.create_sheet(sheet_name)
        for j, col in enumerate(df.columns, start=1):
            s.cell(row=1, column=j, value=col)
        header_row(s, 1, len(df.columns))
        for i, (_, row) in enumerate(df.iterrows(), start=2):
            for j, col in enumerate(df.columns, start=1):
                cell = s.cell(row=i, column=j, value=row[col])
                cell.font = Font(name=FONT, size=10)
                cell.alignment = LFT if j == 1 else CTR
                cell.border = THIN
        s.freeze_panes = "A2"
        autosize(s, len(df.columns), len(df) + 1)

    wb.save(OUT_PATH)
    print(f"Dashboard saved: {OUT_PATH}")


if __name__ == "__main__":
    main()
